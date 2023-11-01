import openpyxl as op


def check_inn_format(file):
    incorrect_int_format = {}
    incorrect_inn = {}
    delite_rows = {}
    try:
        wb = op.load_workbook(file)
        ws = wb['Раздел 1']
        for row in ws.iter_cols(min_col=4, max_col=4, min_row=7):
            for cell in row:
                if cell.value is None:
                    number_of_rows = cell.row
                    print(number_of_rows)
                    empty_region = ws[f'B{number_of_rows}'].value
                    empty_name = ws[f'C{number_of_rows}'].value
                    if (empty_region and empty_name) is None:
                        delite_rows[number_of_rows] = cell.value
                        ws.delete_rows(number_of_rows)
                    else:
                        pass
                else:
                    coordinate_of_cells = cell.coordinate
                    try:
                        int(cell.value)
                    except ValueError:
                        incorrect_int_format[coordinate_of_cells] = cell.value
                    inn_len = len(str(cell.value))
                    if inn_len != 10 and inn_len != 12:
                        incorrect_inn[coordinate_of_cells] = cell.value
        wb.save(file)
    except Exception as error:
        print(error)
    return incorrect_inn


print(check_inn_format(r'C:\ФРПВ_1 раздел\source_data\0100.xlsx'))
