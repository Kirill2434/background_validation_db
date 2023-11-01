import glob
import os
from pathlib import Path
from pprint import pprint

import pandas as pd
from pandas import ExcelWriter
import openpyxl as op
from prettytable import PrettyTable
from tqdm import tqdm

from FRPV.frpv_config import FRPV_CHECK_REPORT_N, FRPV_CHECK_REPORT, R_1, R_3
from source.config import main_folder


def head_of_table(file, sheet_name: str = 'Лист1'):
    """

    @param file:
    @param sheet_name:
    @return:
    """
    miss_files = []
    head_of_file = []
    none_list = []
    num_of_col = 1
    index_of_head_row = 0
    count_of_none = 0
    df = pd.ExcelFile(file)
    file_name = Path(df).name
    wb = op.load_workbook(file)

    try:
        ws = wb[sheet_name]
    except KeyError:
        ws = wb.active

    while True:
        for col in ws.iter_cols(min_col=num_of_col, max_col=num_of_col, max_row=20):
            for cell in col:
                none_list.append(cell.value)
        if list(set(none_list))[0] is None:
            num_of_col += 1
            continue
        break

    for row in ws.iter_cols(min_col=num_of_col, max_col=num_of_col):
        for cell in row:
            if cell.value is None:
                pass
            else:
                try:
                    int_cell = int(cell.value)
                    if int_cell == 1:
                        # print(int_cell)
                        num_of_col += 1
                        # print(num_of_col)
                        for row_second in ws.iter_cols(min_col=num_of_col, max_col=num_of_col):
                            for cell_second in row_second:
                                if cell_second.value is None:
                                    pass
                                else:
                                    try:
                                        int_cell = int(cell_second.value)
                                        str_cell = str(cell_second.value)
                                        # print(int_cell)
                                        # print(str_cell)
                                        if len(str_cell) == 1:
                                            if int_cell == 2:

                                                number_of_rows = cell.row
                                                number_of_cols = cell.column
                                                index_of_head_row += number_of_rows
                                            else:
                                                break
                                        else:
                                            pass
                                    except ValueError:
                                        pass
                except ValueError:
                    pass

    for row in ws.iter_rows(min_row=index_of_head_row, max_row=index_of_head_row):
        try:
            for cell in row:
                if cell.value is None:
                    pass
                else:
                    head_of_file.append(int(cell.value))
        except ValueError:
            miss_files.append(file_name)
            return miss_files
    return head_of_file


# print(head_of_table(r'C:\Приложения регионов_исход\3_14.03.2023 после 11\7200.xlsx'))


def proverka_kolonok_ultimate_xxx(path, sum_of_head: int, *, sheet_name: str = 'Лист1'):
    inc_dict = {}
    final_dict = {}
    sum_of_head = list(range(1, sum_of_head + 1))
    try:
        for file in tqdm(path):
            df = pd.ExcelFile(file)
            file_name = Path(df).name
            head_func = head_of_table(file, sheet_name)
            if head_func == sum_of_head:
                pass
            else:
                final_dict[file_name] = sheet_name
    except FileNotFoundError:
        return 'Файл не найден!'
    return final_dict


# all_filas = glob.glob(r'C:\Приложения регионов_исход\3_14.03.2023 после 11\*.xlsx')
# reault = proverka_kolonok_ultimate_xxx(all_filas, 36)
# print(reault)


def count_empty_rows():
    """Подсчет пустых строк в таблице файлов Excel. """
    pass


def count_files():
    """Финальный подсчет обрабатываемых файлов в рабочей папке. """
    xlsx_files = f"В папке {main_folder} хранится {len(list(glob.glob('*.xlsx')))} объектов в формате .xlsx\n"
    xls_files = f"В папке {main_folder} хранится {len(list(glob.glob('*.xls')))} объектов в формате .xls"
    xlsx_and_xls = (xlsx_files, xls_files)
    return xlsx_and_xls


def record_to_excel(obj, sheet_name):
    """Универсальная функция записи результатов выполнения различных
    проверок и действий в других модулях проекта.

    @param obj: датафрейм который необходимо записать в файл формата xlsx
    @param sheet_name: имя листа на который будут записаны данные
    """

    # если файл-отчет отсувует в папке, то записываем первый файл-отчет
    if os.path.exists(FRPV_CHECK_REPORT) is False:
        with ExcelWriter(FRPV_CHECK_REPORT,
                         engine='openpyxl') as writer:
            # проверяем тип объекта
            # либо список
            if isinstance(obj, list):
                df = pd.DataFrame(obj).transpose()
                df.to_excel(writer, sheet_name=sheet_name)
            # либо словарь
            if isinstance(obj, dict):
                df = pd.DataFrame.from_dict(obj)
                df.to_excel(writer, sheet_name=sheet_name)
    # когда файл уже есть:
    else:
        # счиаем кол-во файлов-отчетов в папке
        sum_of_files = len(list(glob.glob(r'C:\generation_results\check_report_file??.xlsx')))
        num = sum_of_files + 1
        # создаем новое имя файла с учетом нового номера
        new_file_name = FRPV_CHECK_REPORT_N + '_' + str(num) + '.xlsx'
        with ExcelWriter(new_file_name,
                         engine='openpyxl') as writer:
            # проверяем тип объекта
            # либо список
            if isinstance(obj, list):
                df = pd.DataFrame(obj).transpose()
                df.to_excel(writer, sheet_name=sheet_name)
            # либо словарь
            if isinstance(obj, dict):
                df = pd.DataFrame.from_dict(obj)
                df.to_excel(writer, sheet_name=sheet_name)
